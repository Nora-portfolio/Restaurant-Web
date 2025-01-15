from flask import Flask, request, jsonify
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

# Define the path to the Excel file
EXCEL_FILE = 'contact_data.xlsx'

# Ensure the Excel file exists
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Full Name", "Email", "Phone Number", "Subject", "Message"])
    wb.save(EXCEL_FILE)

@app.route('/submit', methods=['POST'])
def submit_form():
    data = request.form
    full_name = data.get('full_name')
    email = data.get('email')
    phone_number = data.get('phone_number')
    subject = data.get('subject')
    message = data.get('message')

    # Open the Excel file and add data
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([full_name, email, phone_number, subject, message])
    wb.save(EXCEL_FILE)

    return jsonify({"message": "Data successfully saved!"}), 200

if __name__ == '__main__':
    app.run(debug=True)

